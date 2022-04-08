
from openpyxl import Workbook, load_workbook


class adminLogin:
    AdminUsername = "Admin"
    AdminPassword = "123456"

    def login_admin(self):
        adminusername=input("Enter Your Username: ")
        adminpassword=input("Enter Your Password: ")
        if adminpassword==self.AdminPassword and adminusername==self.AdminUsername:
            return True
        else:
            return False

    def addMovie(self):
        print("*** Welcome Admin ***")
        wb=load_workbook("MovieData.xlsx")
        ws=wb.active
        rows=len(list(ws.rows))
        number =int(input("How many movies you want to add: "))
        for i in range(rows,number+rows):
             Title=input("Enter Title: ")
             Genre=input("Enter Genre: ")
             Length=input("Enter Length: ")
             Cast=input("Enter Cast: ")
             Capacity=input("Enter Capacity: ")
             AdminRatings=input("Enter Admin Ratings out of 10: ")
             ws.append([Title, Genre, Length, Cast, Capacity, AdminRatings])
        wb.save("MovieData.xlsx")

    def EditMovie(self):
        print("*** Welcome Admin ***")
        wb=load_workbook("MovieData.xlsx")
        ws=wb.active
        rows=len(list(ws.rows))
        edit=input("Which movie you want to Edit: ")
        range=ws["A1":"A"+str(rows)]
        row_number=0
        for movie in range:
            for x in movie:
                row_number+=1
                if x.value==edit:
                    Title = input("Enter New title: ")
                    ws["A"+str(row_number)]=Title
                    Genre = input("enter new genre: ")
                    ws["B"+str(row_number)]=Genre
                    Length = input("Enter new length: ")
                    ws["C"+str(row_number)]=Length
                    Cast = input("Enter new cast: ")
                    ws["D"+str(row_number)]=Cast
                    Capacity=input("Enter new capacity: ")
                    ws["E"+str(row_number)]=Capacity
                    Adminrating=input("Enter new admin rating: ")
                    ws["F"+str(row_number)]=Adminrating
                    break
        wb.save("MovieData.xlsx")

    def deleteMovie(self):
        print("*** Welcome Admin ***")
        wb = load_workbook("MovieData.xlsx")
        ws = wb.active
        rows = len(list(ws.rows))
        delete = input("Which movie you want to delete: ")
        range = ws["A1":"A" + str(rows)]
        row_number = 0
        for movie in range:
            for x in movie:
                row_number += 1
                if x.value == delete:
                    ws.delete_rows(row_number)
                    break
        wb.save("MovieData.xlsx")

