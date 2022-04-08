from openpyxl import load_workbook


class register:
    def userregister(self):
            username = input('Enter username: ')
            if username in open('register.txt', 'r').read():
                print('Username already exists')
                return False
            password = input('Enter password: ')
            c_password = input('Enter confirm password: ')
            if password != c_password:
                print('password did not match')
                return False
            file = open('register.txt', 'a')
            file.write(username)
            file.write(' ')
            file.write(password)
            file.write('\n')
            file.close()
            print('You are successfully registered ')
            return True

    def userlogin(self):
        print("****Login to BookMyShow: ****")
        username = input('Enter registered username: ')
        password = input('Enter registered password: ')
        get_data = open('register.txt', 'r').readlines()
        users_data = []
        for user in get_data:
            users_data.append(user.split())

        total_user = len(users_data)
        increment = 0
        login_success = 0
        while increment < total_user:
            usernames = users_data[increment][0]
            passwords = users_data[increment][1]
            if username == usernames and password == passwords:
                login_success = 1
            increment += 1

        if login_success == 1:
            print("Login Successful Welcome",username)
            return True
        else:
            print('invalid username or password ')
            return False

    def movies(self):
        id=1
        print("*** Movies List ***")
        wb = load_workbook("MovieData.xlsx")
        ws = wb.active
        rows = len(list(ws.rows))
        range = ws["A2":"A" + str(rows)]
        for movie in range:
            for x in movie:
                id+=1
                print(id, " ", x.value)
    def movieInfo(self,choice):
        wb=load_workbook("MovieData.xlsx")
        ws=wb.active
        range = ws[str(choice)]
        print("*** Movie Information ***")
        print("||Title||    ||Genre||    ||length||     ||Cast||      ||Capacity||     ||AdminRatings||")
        for movies in range:
            print(movies.value,end="     ")
        wb.save("MovieData.xlsx")
        print()
    def cancelTickets(self,choice):
        wb = load_workbook("MovieData.xlsx")
        ws = wb.active
        cancel=int(input("Number of seats you want to cancel: "))
        ws["E"+str(choice)]=int(ws["E"+str(choice)].value)-cancel
        wb.save("MovieData.xlsx")
    def Userratings(self,choice):
        wb = load_workbook("MovieData.xlsx")
        ws = wb.active
        user_rating=input("Enter ratings: ")
        ws["G" + str(choice)] = user_rating
        wb.save("MovieData.xlsx")



